import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
from collections import defaultdict
import openpyxl
import re
from dataclasses import dataclass, field
import random

@dataclass
class ProjectData:
    name: str
    note: str = ""
    cells: list = field(default_factory=list)

class AssignmentApp:
    def __init__(self, master):
        self.master = master
        self.master.title("地點導向工作分配系統")
        self.staff = []
        self.projects_by_location = defaultdict(list)
        self.locations = []
        self.assignments = defaultdict(list)
        self.assignment_widgets = {}
        self.file_paths = {}
        self.tabs = {}
        self.setup_ui()

    def setup_ui(self):
        frame = ttk.Frame(self.master)
        frame.pack(padx=10, pady=10, fill='x')
        ttk.Button(frame, text="載入人員", command=self.load_staff).grid(row=0, column=0, padx=5)
        ttk.Button(frame, text="載入項目", command=self.load_projects).grid(row=0, column=1, padx=5)
        ttk.Button(frame, text="載入地點", command=self.load_locations).grid(row=0, column=2, padx=5)
        ttk.Button(frame, text="開始分配", command=self.start_assignment).grid(row=1, column=0, pady=10)
        ttk.Button(frame, text="匯出結果", command=self.export_results).grid(row=1, column=1, pady=10)
        ttk.Button(frame, text="測試自動分配", command=self.test_auto_assign).grid(row=1, column=2, pady=10)
        self.notebook = ttk.Notebook(self.master)
        self.notebook.pack(expand=True, fill='both')

    def load_staff(self):
        path = filedialog.askopenfilename(title="選擇人員.csv")
        if not path:
            return
        with open(path, encoding="utf-8") as f:
            self.staff = [line.strip() for line in f if line.strip()]
        messagebox.showinfo("完成", f"載入 {len(self.staff)} 位人員")

    def load_projects(self):
        path = filedialog.askopenfilename(title="選擇項目.csv")
        if not path:
            return
        self.projects_by_location.clear()
        with open(path, encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) < 2:
                    continue
                loc = row[0].strip()
                name = row[1].strip()
                note = row[2].strip() if len(row) >= 3 else ""
                cells = [c.strip() for c in row[3].split("|")] if len(row) >= 4 and row[3].strip() else []
                self.projects_by_location[loc].append(ProjectData(name, note, cells))
        messagebox.showinfo("完成", "已載入項目")

    def load_locations(self):
        path = filedialog.askopenfilename(title="選擇地點.csv")
        if not path:
            return
        with open(path, encoding="utf-8") as f:
            self.locations = [line.strip() for line in f if line.strip()]
        messagebox.showinfo("完成", f"載入 {len(self.locations)} 個地點")

    def start_assignment(self):
        from tkinter.simpledialog import askstring
        for tab in self.notebook.tabs():
            self.notebook.forget(tab)
        self.assignment_widgets.clear()
        self.assignments.clear()
        for loc in self.locations:
            canvas = tk.Canvas(self.notebook)
            frame = ttk.Frame(canvas)
            vsb = ttk.Scrollbar(canvas, orient='vertical', command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            canvas.pack(side='left', fill='both', expand=True)
            vsb.pack(side='right', fill='y')
            canvas.create_window((0,0), window=frame, anchor='nw')
            frame.bind('<Configure>', lambda e, c=canvas: c.configure(scrollregion=c.bbox('all')))
            canvas.bind_all('<MouseWheel>', lambda e, c=canvas: c.yview_scroll(int(-1*(e.delta/120)), 'units'))
            self.notebook.add(canvas, text=loc)
            self.assignment_widgets[loc] = []
            projects = self.projects_by_location.get(loc, [])
            rows, cols = 10, 6
            grid = ttk.Frame(frame)
            grid.pack(fill='both', expand=True)
            for idx, project in enumerate(projects):
                r, c = divmod(idx, cols)
                cell = ttk.Frame(grid, padding=2)
                cell.grid(row=r, column=c, sticky='w')
                ttk.Label(cell, text=f"{project.name}（{project.note}）" if project.note else project.name,
                          wraplength=120, justify='left').pack(anchor='w')
                var = tk.StringVar()
                cb = ttk.Combobox(cell, textvariable=var, values=self.get_available_staff(loc), state='readonly')
                cb.pack(fill='x')
                cb.bind('<<ComboboxSelected>>', lambda e, l=loc: self.update_assignments(l))
                ttk.Button(cell, text='清除', command=lambda v=var, l=loc: self.clear_assignment(v,l)).pack(pady=1)
                self.assignment_widgets[loc].append((project, var, cb))
            if loc == '2樓禪堂':
                def add_proj():
                    opts = [p.name for p in self.projects_by_location[loc]]
                    name = askstring('新增項目', f'請選擇：{opts}')
                    if name in opts:
                        project = next(p for p in self.projects_by_location[loc] if p.name == name)
                        idx = len(self.assignment_widgets[loc])
                        r, c = divmod(idx, cols)
                        cell = ttk.Frame(grid, padding=2); cell.grid(row=r, column=c, sticky='w')
                        ttk.Label(cell, text=project.name, wraplength=120, justify='left').pack(anchor='w')
                        var = tk.StringVar()
                        cb = ttk.Combobox(cell, textvariable=var, values=self.get_available_staff(loc), state='readonly')
                        cb.pack(fill='x')
                        cb.bind('<<ComboboxSelected>>', lambda e, l=loc: self.update_assignments(l))
                        ttk.Button(cell, text='清除', command=lambda v=var, l=loc: self.clear_assignment(v,l)).pack(pady=1)
                        self.assignment_widgets[loc].append((project, var, cb))
                ttk.Button(frame, text='➕ 新增項目', command=add_proj).pack(pady=5)
            canvas.yview_moveto(0)

    def get_available_staff(self, loc):
        used = {s for l in self.assignments if l != loc for _, s in self.assignments[l]}
        return [s for s in self.staff if s not in used]

    def clear_assignment(self, var, loc):
        var.set('')
        self.update_assignments(loc)

    def test_auto_assign(self):
        if not self.staff or not self.assignment_widgets:
            messagebox.showerror('錯誤','請先載入人員與項目並開始分配')
            return
        ppl = self.staff[:]; random.shuffle(ppl)
        locs = list(self.assignment_widgets.keys())
        per = len(ppl)//len(locs); rem = len(ppl)%len(locs); i = 0
        for idx, l in enumerate(locs):
            cnt = per + (1 if idx < rem else 0)
            pool = ppl[i:i+cnt]; i += cnt
            for proj, var, _ in self.assignment_widgets[l]:
                if not var.get() and pool:
                    var.set(pool.pop(0))
            filled = [var.get() for _, var, _ in self.assignment_widgets[l] if var.get()]
            for proj, var, _ in self.assignment_widgets[l]:
                if not var.get() and filled:
                    var.set(random.choice(filled))
            self.update_assignments(l)

    def update_assignments(self, loc):
        self.assignments[loc] = []
        for proj, var, _ in self.assignment_widgets[loc]:
            if var.get(): self.assignments[loc].append((proj, var.get()))
        for proj, var, cb in self.assignment_widgets[loc]:
            vals = self.get_available_staff(loc)
            cb['values'] = vals
            if var.get() not in vals: var.set('')

    def generate_cell_content(self):
        m = defaultdict(list)
        for loc, entries in self.assignments.items():
            for proj, staff in entries:
                for cell in proj.cells:
                    if cell == 'D31':
                        val = f"【{loc}】{proj.name}：{staff}"
                    else:
                        val = f"{proj.name}：{staff}"
                    m[cell].append(val)
        return m

    def write_to_excel(self, wb, cell_content):
        ws = wb.active
        for cell, lines in cell_content.items():
            ws[cell] = "\n".join(lines)
            ws[cell].alignment = openpyxl.styles.Alignment(wrap_text=True)
            ws.row_dimensions[ws[cell].row].height = 15 * len(lines)

    def export_results(self):
        try:
            import pandas as pd
        except ImportError:
            messagebox.showerror('缺少函式庫','請安裝 pandas: pip install pandas')
            return
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror('缺少函式庫','請安裝 openpyxl: pip install openpyxl')
            return
        path = filedialog.askopenfilename(title='選擇 Excel 模板',filetypes=[('Excel','*.xlsx')])
        if not path: return
        from shutil import copyfile
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        out_path = path.replace('.xlsx', f'_分配結果_{timestamp}.xlsx')
        copyfile(path, out_path)
        wb = openpyxl.load_workbook(out_path)
        map_path = filedialog.askopenfilename(title='選擇 項目對應 CSV',filetypes=[('CSV','*.csv')])
        if not map_path: return
        df = pd.read_csv(map_path)
        content = defaultdict(list)
        for loc, entries in self.assignments.items():
            for proj, staff in entries:
                rows = df[(df['地點'] == loc) & (df['項目名稱'] == proj.name)]
                for _, r in rows.iterrows():
                    for c in str(r['儲存格']).split('|'):
                        cell = c.strip()
                        if not cell: continue
                        if cell in ['D31', 'D30', 'E10']:
                            label = f"【{loc}】{proj.name}：{staff}"
                        else:
                            label = f"{proj.name}：{staff}"
                        content[cell].append(label)
        for cell, lines in content.items():
            ws = wb.active
            ws[cell] = "\n".join(lines)
            ws[cell].alignment = openpyxl.styles.Alignment(wrap_text=True)
            ws.row_dimensions[ws[cell].row].height = 15 * len(lines)
        save_path = filedialog.asksaveasfilename(defaultextension='.xlsx',filetypes=[('Excel','*.xlsx')])
        if save_path:
            wb.save(save_path)
            messagebox.showinfo('完成', f'已儲存：{save_path}')

if __name__ == '__main__':
    root = tk.Tk()
    AssignmentApp(root)
    root.mainloop()
